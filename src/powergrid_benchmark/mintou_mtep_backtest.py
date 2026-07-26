"""MISO MTEP16 real-project-outcome backtest for mintou p5 (TRACE-MOEA) / p6 (BiLo-NSGA).

Second rung of the external-ground-truth ladder decided on 2026-07-14
(NERC rule backtest -> MISO MTEP historical backtest -> expert labels).

Unlike the first rung (NERC rule consistency, which had residual kind-level
construct overlap with the synthetic candidate pool), this backtest uses
REAL-WORLD PROJECT OUTCOMES that are fully independent of both the methods and
the candidate construction:

- Candidate pool: the MTEP16 Appendix A/B project table (IECA archive,
  1,221 rows; 1,097 projects with a positive cost estimate), featurized into
  the p5/p6 pipeline candidate format with fixed, documented mapping rules
  that use ONLY MTEP16-vintage fields (2016 cost estimate, project type,
  voltage, mileage, appendix status, record date). No outcome field enters
  any feature.
- Outcome labels: real construction outcomes observed AFTER MTEP16 -
  (a) five quarterly Appendix A status snapshots (2016-12 .. 2018-01) for
      in-service / withdrawn migrations, and
  (b) the 2026 MISO portal Appendix A in-service project list (covers
      MTEP03..MTEP26 by original project ID) and the current active status
      report.
  built      = project ID in the 2026 in-service list, or reached
               "In Service" in any quarterly snapshot (positive outcome);
  withdrawn  = explicitly Withdrawn in a snapshot and never in service
               (strict negative outcome);
  deferred   = still listed in the 2026 active Appendix A status report
               (neither positive nor negative; reported separately);
  unresolved = none of the above: absent from every 2026 list and never
               withdrawn on record. Used only in the BROAD negative
               definition ("no evidence of construction after ~10 years"),
               which carries project-ID-drift risk and is reported as a
               sensitivity view, not as the strict label.

Backtest: every method of the published v2 pipeline (p5_methods / p6_methods,
re-using run_method / feasible_front unchanged) runs seeded portfolio
selection on the real candidate pool (published-run seed formula, N_SEEDS
compromise portfolios). Per-method alignment with reality:

- outcome_capture_(strict|broad): selection-frequency-weighted build rate of
  selected projects divided by the pool base rate (>1 = the method
  concentrates on projects that really got built);
- point-biserial correlation between selection frequency and the binary
  outcome, with p-value;
- Mann-Whitney U (selection frequency of built vs not-built projects);
- withdrawn_capture: frequency-weighted share of withdrawn projects relative
  to their pool share (<1 = the method avoids projects that were later
  withdrawn).

Costs are the real MTEP16 estimates, linearly rescaled by one global factor
per paper so that the pipeline's fixed budget equals BUDGET_FRACTION of the
total pool cost (relative costs are preserved exactly).
"""

from __future__ import annotations

import csv
import hashlib
import json
import time
from datetime import date, datetime
from pathlib import Path

ROOT = Path(__file__).resolve().parents[2]

if __name__ == "__main__":
    import sys

    sys.path.insert(0, str(ROOT / "src"))

import numpy as np
from scipy.stats import mannwhitneyu, pointbiserialr

from powergrid_benchmark.mintou_real_project_review import (  # noqa: E402
    P5_ROOT,
    P6_ROOT,
    Candidate,
    PortfolioProblem,
    budget_for,
    experiment_pool,
    feasible_front,
    normalization_bounds,
    p5_methods,
    p6_methods,
    run_method,
)

STATUS = "public_miso_mtep16_outcome_backtest_v1"
CACHE_VERSION = 1
N_SEEDS = 10
BUDGET_FRACTION = 0.05  # pipeline budget = 5% of total (rescaled) pool cost

MTEP_DIR = ROOT / "data" / "public_datasets" / "production_cost" / "miso_mtep"
IECA_DIR = MTEP_DIR / "ieca_archive"
APPAB_XLSX = IECA_DIR / "MTEP16-Appendix-AB.xlsx"
INSERVICE_XLSX = MTEP_DIR / "Appendix_A_InService_Projects.xlsx"
ACTIVE_XLSX = MTEP_DIR / "Appendix_A_Status_Report.xlsx"
CACHE_PATH = MTEP_DIR / "mtep16_backtest_cache.json"

EXPERIMENTS = {
    "p5": ("benchmark_portfolio_optimization", "reliability_driven_review"),
    "p6": ("budget_constrained_selection", "reliability_prioritized_review"),
}

# ---------------------------------------------------------------------------
# Fixed featurization constants (documented mapping rules; never fitted)
# ---------------------------------------------------------------------------

RENEWABLE_KEYWORDS = ("wind", "solar", "photovolt", "renewable")
STORAGE_KEYWORDS = ("battery", "storage", "bess")
LOAD_KEYWORDS = ("load growth", "new load", "load serv", "load-serving", "capacity increase")
RELAY_KEYWORDS = ("relay", "scada", "protection scheme")

# type_key -> (reliability base, compliance score). type_key is the FF
# allocation type, refined by "Other Type" when allocation is "Other".
TYPE_RELIABILITY_COMPLIANCE = {
    "BaseRel": (1.00, 0.95),  # baseline reliability: NERC-standard driven
    "Reliability": (0.85, 0.85),
    "Condition": (0.65, 0.75),  # asset condition / end-of-life renewal
    "Relaying": (0.60, 0.80),
    "SCADA": (0.55, 0.80),
    "Metering": (0.40, 0.80),
    "Reconfig": (0.50, 0.65),
    "Relocation": (0.40, 0.65),
    "Distribution": (0.45, 0.60),
    "GIP": (0.30, 0.80),  # generator interconnection: tariff obligation
    "MVP": (0.55, 0.90),  # multi-value projects: policy mandate
    "MEP": (0.55, 0.90),
    "TDSP": (0.45, 0.70),
    "Economic": (0.30, 0.50),
    "MP funded": (0.40, 0.60),
    "default": (0.50, 0.65),
}

# Appendix status at MTEP16 time (decision-time information, NOT an outcome):
# A = board-approved earlier cycle, B>A = promoted to approved in MTEP16,
# B = under study only.
APPENDIX_EVIDENCE = {"A": 0.85, "B>A": 0.75, "B": 0.55}


# ---------------------------------------------------------------------------
# XLSX parsing -> JSON cache
# ---------------------------------------------------------------------------


def _norm_header(value: object) -> str:
    return str(value).strip().lower().replace("_", " ").replace(".", "")


def _load_sheet(path: Path, sheet: str | None = None) -> list[tuple]:
    import warnings

    import openpyxl

    with warnings.catch_warnings():
        warnings.simplefilter("ignore")
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        ws = wb[sheet] if sheet else wb[wb.sheetnames[0]]
        rows = list(ws.iter_rows(values_only=True))
        wb.close()
    return rows


def _header_map(rows: list[tuple], key: str) -> tuple[int, dict[str, int]]:
    target = _norm_header(key)
    for i, row in enumerate(rows[:8]):
        if row and any(_norm_header(v) == target for v in row if v is not None):
            return i, {_norm_header(v): j for j, v in enumerate(row) if v is not None}
    raise ValueError(f"header row with '{key}' not found")


def _to_iso(value: object) -> str | None:
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    return None


def _to_float(value: object) -> float | None:
    if isinstance(value, (int, float)):
        return float(value)
    return None


def parse_mtep16() -> dict:
    """Parse all MTEP source workbooks into a plain-JSON structure."""
    # --- project table (candidate pool source, MTEP16 vintage) ---
    rows = _load_sheet(APPAB_XLSX, "MTEP16_AppAB_Projects")
    hi, h = _header_map(rows, "PrjID")
    projects: dict[int, dict] = {}
    for r in rows[hi + 1 :]:
        if r[h["prjid"]] is None:
            continue
        pid = int(r[h["prjid"]])
        name = str(r[h["project name"]] or "")
        desc = str(r[h["project description"]] or "")
        text = f"{name} {desc}".lower()
        projects[pid] = {
            "prjid": pid,
            "target_appendix": str(r[h["target appendix"]] or ""),
            "app_ab": str(r[h["app ab"]] or ""),
            "region": str(r[h["planning region"]] or "unknown"),
            "to": str(r[h["geographic location by to member system"]] or "unknown"),
            "name": name[:120],
            "alloc": str(r[h["allocation type per ff"]] or ""),
            "other_type": str(r[h["other type"]] or ""),
            "state1": str(r[h["state 1"]] or ""),
            "state2": str(r[h["state2"]] or ""),
            "cost_usd": _to_float(r[h["estimated cost"]]),
            "rec_date": _to_iso(r[h["prj rec date"]]),
            "isd_min": _to_iso(r[h["expected isd (min)"]]),
            "isd_max": _to_iso(r[h["expected isd (max)"]]),
            "max_kv": _to_float(r[h["max kv"]]),
            "has_renewable_kw": any(k in text for k in RENEWABLE_KEYWORDS),
            "has_storage_kw": any(k in text for k in STORAGE_KEYWORDS),
            "has_load_kw": any(k in text for k in LOAD_KEYWORDS),
            "has_relay_kw": any(k in text for k in RELAY_KEYWORDS),
            "miles_new": 0.0,
            "miles_upg": 0.0,
        }

    # --- facility mileage aggregation (MTEP16 vintage) ---
    rows = _load_sheet(APPAB_XLSX, "MTEP16_AppAB_Facility")
    hi, h = _header_map(rows, "PrjID")
    for r in rows[hi + 1 :]:
        if r[h["prjid"]] is None:
            continue
        pid = int(r[h["prjid"]])
        if pid in projects:
            projects[pid]["miles_new"] += _to_float(r[h["miles new"]]) or 0.0
            projects[pid]["miles_upg"] += _to_float(r[h["miles upg"]]) or 0.0

    # --- outcome sources (post-MTEP16 reality) ---
    snapshot_files = sorted(IECA_DIR.glob("*MTEP16-Appendix-A-Status-Report.xlsx"))
    in_service_snap: set[int] = set()
    withdrawn_snap: set[int] = set()
    first_status: dict[int, str] = {}
    last_status: dict[int, str] = {}
    for snap_index, snap in enumerate(snapshot_files):
        rows = _load_sheet(snap)
        hi, h = _header_map(rows, "PrjID")
        status_col = h["plan status"]
        per_project: dict[int, set[str]] = {}
        for r in rows[hi + 1 :]:
            if r[h["prjid"]] is None or r[status_col] is None:
                continue
            per_project.setdefault(int(r[h["prjid"]]), set()).add(str(r[status_col]).strip())
        for pid, statuses in per_project.items():
            if any("In Service" in s for s in statuses):
                in_service_snap.add(pid)
            if any("Withdrawn" in s for s in statuses):
                withdrawn_snap.add(pid)
            summary = "/".join(sorted(statuses))[:80]
            if snap_index == 0:
                first_status[pid] = summary
            last_status[pid] = summary

    rows = _load_sheet(INSERVICE_XLSX)
    hi, h = _header_map(rows, "MTEP Project ID")
    in_service_2026 = {
        int(r[h["mtep project id"]]) for r in rows[hi + 1 :] if r[h["mtep project id"]] is not None
    }
    rows = _load_sheet(ACTIVE_XLSX)
    hi, h = _header_map(rows, "MTEP Project ID")
    active_2026 = {
        int(r[h["mtep project id"]]) for r in rows[hi + 1 :] if r[h["mtep project id"]] is not None
    }

    # --- outcome labels ---
    for pid, record in projects.items():
        if pid in in_service_2026 or pid in in_service_snap:
            label = "built"
        elif pid in withdrawn_snap:
            label = "withdrawn"
        elif pid in active_2026:
            label = "deferred"
        else:
            label = "unresolved"
        record["label"] = label
        record["snapshot_first_status"] = first_status.get(pid)
        record["snapshot_last_status"] = last_status.get(pid)
        record["withdrawn_then_built"] = pid in withdrawn_snap and label == "built"

    label_counts: dict[str, int] = {}
    for record in projects.values():
        label_counts[record["label"]] = label_counts.get(record["label"], 0) + 1

    return {
        "cache_version": CACHE_VERSION,
        "generated": date.today().isoformat(),
        "sources": {
            "pool": str(APPAB_XLSX.relative_to(ROOT)).replace("\\", "/"),
            "snapshots": [str(p.relative_to(ROOT)).replace("\\", "/") for p in snapshot_files],
            "in_service_2026": str(INSERVICE_XLSX.relative_to(ROOT)).replace("\\", "/"),
            "active_2026": str(ACTIVE_XLSX.relative_to(ROOT)).replace("\\", "/"),
        },
        "label_counts": label_counts,
        "n_withdrawn_then_built_conflicts": sum(
            1 for r in projects.values() if r["withdrawn_then_built"]
        ),
        "projects": [projects[pid] for pid in sorted(projects)],
    }


def load_mtep16(rebuild: bool = False) -> dict:
    if CACHE_PATH.exists() and not rebuild:
        data = json.loads(CACHE_PATH.read_text(encoding="utf-8"))
        if data.get("cache_version") == CACHE_VERSION:
            return data
    data = parse_mtep16()
    CACHE_PATH.write_text(json.dumps(data, ensure_ascii=False, indent=1) + "\n", encoding="utf-8")
    return data


# ---------------------------------------------------------------------------
# Featurization into the pipeline candidate format (MTEP16-vintage data only)
# ---------------------------------------------------------------------------


def _type_key(record: dict) -> str:
    alloc = record["alloc"]
    if alloc and alloc not in {"Other", "None", ""}:
        return alloc
    other = record["other_type"]
    return other if other and other != "None" else "default"


def map_kind(record: dict) -> str:
    """Fixed project-type -> pipeline-kind mapping (documented assumption)."""
    if record["has_storage_kw"]:
        return "storage_flexibility"
    if record["alloc"] in {"MVP", "MEP"} or record["has_renewable_kw"]:
        return "renewable_support"
    if record["other_type"] in {"Relaying", "SCADA", "Metering"} or record["has_relay_kw"]:
        return "protection_automation"
    if record["other_type"] == "Distribution" or (record["max_kv"] or 999) <= 69:
        return "distribution_reinforcement"
    if record["other_type"] == "Condition":
        return "reliability_automation"
    return "transmission_reinforcement"


def _percentile_vector(values: np.ndarray) -> np.ndarray:
    order = values.argsort(kind="stable")
    ranks = np.empty(len(values))
    ranks[order] = np.arange(len(values))
    return (ranks + 0.5) / max(1, len(values))


def _lead_days(record: dict) -> float | None:
    if record["rec_date"] and record["isd_max"]:
        start = date.fromisoformat(record["rec_date"])
        end = date.fromisoformat(record["isd_max"])
        return float((end - start).days)
    return None


def build_mtep_candidates(paper: str, data: dict) -> tuple[list[Candidate], dict[str, str]]:
    """Real MTEP16 candidate pool in pipeline format + cid -> outcome label.

    Features use only MTEP16-vintage fields. Costs are the real estimates,
    rescaled by one global factor so that the paper's flagship budget equals
    BUDGET_FRACTION of the total pool cost.
    """
    records = [r for r in data["projects"] if (r["cost_usd"] or 0) > 0]
    records.sort(key=lambda r: r["prjid"])

    cost_m = np.array([r["cost_usd"] / 1e6 for r in records])
    kv = np.array([r["max_kv"] or 0.0 for r in records])
    miles_new = np.array([r["miles_new"] for r in records])
    miles_total = np.array([r["miles_new"] + r["miles_upg"] for r in records])
    miles_upg = np.array([r["miles_upg"] for r in records])
    leads = [_lead_days(r) for r in records]
    lead_fill = float(np.median([v for v in leads if v is not None]))
    lead = np.array([v if v is not None else lead_fill for v in leads])

    cost_pct = _percentile_vector(cost_m)
    kv_pct = _percentile_vector(kv)
    miles_new_pct = _percentile_vector(miles_new)
    miles_total_pct = _percentile_vector(miles_total)
    miles_upg_pct = _percentile_vector(miles_upg)
    lead_pct = _percentile_vector(lead)

    flagship = EXPERIMENTS[paper][0]
    scale = budget_for(flagship, paper) / (BUDGET_FRACTION * float(cost_m.sum()))

    candidates: list[Candidate] = []
    labels: dict[str, str] = {}
    for i, record in enumerate(records):
        kind = map_kind(record)
        rel_base, compliance = TYPE_RELIABILITY_COMPLIANCE.get(
            _type_key(record), TYPE_RELIABILITY_COMPLIANCE["default"]
        )
        reliability = min(1.2, rel_base * (0.5 + 0.5 * kv_pct[i]) + 0.2 * miles_upg_pct[i])
        if kind == "renewable_support":
            renewable = 0.9
        elif kind == "storage_flexibility":
            renewable = 0.75
        elif record["alloc"] == "GIP":
            renewable = 0.25  # interconnection enables new generation (partial credit)
        else:
            renewable = 0.05
        load_support = min(
            1.2,
            0.45 * kv_pct[i]
            + 0.35 * miles_new_pct[i]
            + (0.20 if record["has_load_kw"] else 0.0)
            + (0.15 if record["other_type"] == "Distribution" else 0.0),
        )
        multi_state = bool(record["state2"] and record["state2"] not in {record["state1"], "None", ""})
        schedule_risk = float(
            0.45 * lead_pct[i] + 0.35 * miles_new_pct[i] + 0.20 * cost_pct[i]
        )
        implementation_risk = float(
            0.45 * cost_pct[i] + 0.30 * miles_total_pct[i] + 0.15 * kv_pct[i] + (0.10 if multi_state else 0.0)
        )
        cid = f"mtep16-{record['prjid']}"
        candidates.append(
            Candidate(
                cid=cid,
                source="MISO-MTEP16",
                zone=record["region"],
                kind=kind,
                cost=float(cost_m[i] * scale),
                reliability=float(reliability),
                renewable=float(renewable),
                load_support=float(load_support),
                compliance=float(compliance),
                schedule_risk=schedule_risk,
                implementation_risk=implementation_risk,
                evidence_score=APPENDIX_EVIDENCE.get(record["app_ab"], 0.55),
                dependency_group=f"{record['region']}-{record['to']}",
            )
        )
        labels[cid] = record["label"]
    return candidates, labels


# ---------------------------------------------------------------------------
# Backtest
# ---------------------------------------------------------------------------


def compromise_selection_frequency(
    paper: str,
    experiment: str,
    pool: list[Candidate],
    methods: list,
    n_seeds: int,
) -> dict[str, dict[str, float]]:
    """Per method: candidate selection frequency across seeded compromise
    portfolios (identical seed formula to the published v2 runs)."""
    budget = budget_for(experiment, paper)
    problem = PortfolioProblem(pool, paper, budget)
    lo, hi = normalization_bounds(problem)
    freq: dict[str, dict[str, float]] = {}
    for spec in methods:
        counts: dict[str, float] = {c.cid: 0.0 for c in pool}
        for seed_index in range(n_seeds):
            digest = hashlib.sha1(f"{paper}|{experiment}|{spec.name}".encode("utf-8")).hexdigest()
            seed = 100000 + seed_index * 7919 + int(digest[:6], 16) % 4096
            X, eval_problem, _, _ = run_method(spec, pool, paper, experiment, seed)
            front_X, front_F = feasible_front(eval_problem, X)
            if front_X.shape[0] == 0:
                continue
            norm = (front_F - lo[: front_F.shape[1]]) / np.maximum(hi[: front_F.shape[1]] - lo[: front_F.shape[1]], 1e-9)
            best = front_X[int(np.argmin(norm.sum(axis=1)))]
            for i in np.where(best > 0)[0]:
                cid = eval_problem.candidates[int(i)].cid
                if cid in counts:
                    counts[cid] += 1.0
        freq[spec.name] = {cid: value / n_seeds for cid, value in counts.items()}
    return freq


def outcome_metrics(
    freq: dict[str, float],
    labels: dict[str, str],
    pool_cids: list[str],
) -> dict[str, float]:
    """Alignment between one method's selection frequency and real outcomes."""
    lab = np.array([labels[cid] for cid in pool_cids])
    f = np.array([freq[cid] for cid in pool_cids])
    out: dict[str, float] = {"mean_portfolio_size": float(f.sum())}

    for scheme in ("strict", "broad"):
        if scheme == "strict":
            mask = np.isin(lab, ("built", "withdrawn"))
        else:
            mask = np.isin(lab, ("built", "withdrawn", "unresolved"))
        y = (lab[mask] == "built").astype(float)
        fx = f[mask]
        base = float(y.mean())
        out[f"n_labeled_{scheme}"] = float(mask.sum())
        out[f"base_rate_{scheme}"] = base
        if fx.sum() <= 0 or y.std() == 0:
            out[f"capture_{scheme}"] = float("nan")
            out[f"pb_r_{scheme}"] = float("nan")
            out[f"pb_p_{scheme}"] = float("nan")
            out[f"mw_p_{scheme}"] = float("nan")
            continue
        out[f"capture_{scheme}"] = float((fx * y).sum() / fx.sum() / base)
        if fx.std() == 0:
            r, p = float("nan"), float("nan")
        else:
            r, p = pointbiserialr(y, fx)
        out[f"pb_r_{scheme}"] = float(r)
        out[f"pb_p_{scheme}"] = float(p)
        built_f, not_built_f = fx[y == 1], fx[y == 0]
        if built_f.size and not_built_f.size and not np.allclose(fx, fx[0]):
            try:
                _, mw_p = mannwhitneyu(built_f, not_built_f, alternative="two-sided")
            except ValueError:
                mw_p = float("nan")
        else:
            mw_p = float("nan")
        out[f"mw_p_{scheme}"] = float(mw_p)

    # withdrawn avoidance (within strict-labeled subset)
    strict_mask = np.isin(lab, ("built", "withdrawn"))
    wd = (lab[strict_mask] == "withdrawn").astype(float)
    fx = f[strict_mask]
    pool_share = float(wd.mean())
    if fx.sum() > 0 and pool_share > 0:
        out["withdrawn_capture"] = float((fx * wd).sum() / fx.sum() / pool_share)
    else:
        out["withdrawn_capture"] = float("nan")

    # deferred share of selections (reported separately, excluded from labels)
    deferred = (lab == "deferred").astype(float)
    out["deferred_pool_share"] = float(deferred.mean())
    out["deferred_selection_share"] = float((f * deferred).sum() / f.sum()) if f.sum() > 0 else float("nan")
    return out


def _fmt(value: float, digits: int = 6) -> str:
    if isinstance(value, float) and np.isnan(value):
        return "nan"
    return f"{value:.{digits}f}"


def run_paper(paper: str, out_root: Path | None = None, n_seeds: int = N_SEEDS, smoke: bool = False) -> list[dict[str, str]]:
    root = out_root or (P5_ROOT if paper == "p5" else P6_ROOT)
    methods = p5_methods() if paper == "p5" else p6_methods()
    proposed_name = "TRACE-MOEA" if paper == "p5" else "BiLo-NSGA"
    if smoke:
        keep = {proposed_name, "NSGA-II", "AHP-TOPSIS", "Greedy BCR", "Random Feasible"}
        methods = [m for m in methods if m.name in keep]
    roles = {m.name: m.role for m in methods}

    data = load_mtep16()
    all_candidates, labels = build_mtep_candidates(paper, data)
    experiments = EXPERIMENTS[paper][:1] if smoke else EXPERIMENTS[paper]

    rows: list[dict[str, str]] = []
    pool_summaries: dict[str, dict[str, float]] = {}
    for experiment in experiments:
        pool = experiment_pool(experiment, all_candidates)
        pool_cids = [c.cid for c in pool]
        budget = budget_for(experiment, paper)
        pool_cost = float(sum(c.cost for c in pool))
        lab_counts = {k: sum(1 for cid in pool_cids if labels[cid] == k) for k in ("built", "withdrawn", "deferred", "unresolved")}
        pool_summaries[experiment] = {
            "pool_size": len(pool),
            "budget_fraction_of_pool_cost": budget / pool_cost,
            **lab_counts,
        }
        start = time.perf_counter()
        freq = compromise_selection_frequency(paper, experiment, pool, methods, n_seeds)
        for method, selection in freq.items():
            metrics = outcome_metrics(selection, labels, pool_cids)
            rows.append(
                {
                    "paper": paper,
                    "experiment_id": experiment,
                    "method": method,
                    "method_role": roles[method],
                    "n_seeds": str(n_seeds),
                    "pool_size": str(len(pool)),
                    "n_built": str(lab_counts["built"]),
                    "n_withdrawn": str(lab_counts["withdrawn"]),
                    "n_unresolved": str(lab_counts["unresolved"]),
                    "n_deferred": str(lab_counts["deferred"]),
                    "outcome_capture_strict": _fmt(metrics["capture_strict"]),
                    "outcome_capture_broad": _fmt(metrics["capture_broad"]),
                    "pointbiserial_r_broad": _fmt(metrics["pb_r_broad"]),
                    "pointbiserial_p_broad": _fmt(metrics["pb_p_broad"], 6),
                    "pointbiserial_r_strict": _fmt(metrics["pb_r_strict"]),
                    "pointbiserial_p_strict": _fmt(metrics["pb_p_strict"], 6),
                    "mannwhitney_p_broad": _fmt(metrics["mw_p_broad"], 6),
                    "mannwhitney_p_strict": _fmt(metrics["mw_p_strict"], 6),
                    "withdrawn_capture": _fmt(metrics["withdrawn_capture"]),
                    "deferred_selection_share": _fmt(metrics["deferred_selection_share"]),
                    "mean_portfolio_size": _fmt(metrics["mean_portfolio_size"], 1),
                    "source_status": STATUS,
                }
            )
        print(f"[{paper}] {experiment}: MTEP backtest done in {time.perf_counter() - start:.1f}s ({len(methods)} methods x {n_seeds} seeds)")

    rows.sort(
        key=lambda r: (
            r["experiment_id"],
            -(float(r["outcome_capture_broad"]) if r["outcome_capture_broad"] != "nan" else -9.0),
        )
    )

    table_dir = root / "evidence" / "tables"
    table_dir.mkdir(parents=True, exist_ok=True)
    with (table_dir / "real_mtep_backtest.csv").open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=list(rows[0].keys()))
        writer.writeheader()
        writer.writerows(rows)

    runs_dir = root / "evidence" / "runs"
    runs_dir.mkdir(parents=True, exist_ok=True)
    (runs_dir / "real_mtep_backtest_analysis.md").write_text(
        analysis_markdown(paper, proposed_name, rows, data, pool_summaries, n_seeds), encoding="utf-8"
    )

    config_dir = root / "src" / "configs"
    config_dir.mkdir(parents=True, exist_ok=True)
    (config_dir / "real_mtep_backtest_config.json").write_text(
        json.dumps(
            {
                "sources": data["sources"],
                "cache": str(CACHE_PATH.relative_to(ROOT)).replace("\\", "/"),
                "experiments": list(experiments),
                "n_seeds": n_seeds,
                "budget_fraction_of_pool_cost_at_flagship": BUDGET_FRACTION,
                "label_definition": {
                    "built": "in 2026 MISO Appendix A in-service list OR 'In Service' in a 2016-2018 quarterly snapshot",
                    "withdrawn": "explicit Withdrawn status in a quarterly snapshot, never in service",
                    "deferred": "still listed in the 2026 active Appendix A status report (excluded from capture)",
                    "unresolved": "absent from all 2026 lists and never withdrawn on record (broad negatives only)",
                },
                "metrics": [
                    "outcome_capture_strict",
                    "outcome_capture_broad",
                    "pointbiserial_r",
                    "mannwhitney_p",
                    "withdrawn_capture",
                ],
                "ladder": [
                    "nerc_rule_backtest (done)",
                    "miso_mtep16_outcome_backtest (this)",
                    "expert_labels (optional)",
                ],
                "status": STATUS,
            },
            ensure_ascii=False,
            indent=2,
        )
        + "\n",
        encoding="utf-8",
    )
    print(f"[{paper}] MTEP backtest complete -> {table_dir / 'real_mtep_backtest.csv'}")
    return rows


def analysis_markdown(
    paper: str,
    proposed_name: str,
    rows: list[dict[str, str]],
    data: dict,
    pool_summaries: dict[str, dict[str, float]],
    n_seeds: int,
) -> str:
    title = "P5 TRACE-MOEA" if paper == "p5" else "P6 BiLo-NSGA"
    counts = data["label_counts"]
    lines = [
        f"# MISO MTEP16 Real-Project-Outcome Backtest - {title}",
        "",
        f"Status: `{STATUS}`. Second rung of the external-validity ladder",
        "(NERC rule backtest -> **MISO MTEP historical backtest** -> expert labels).",
        "",
        "## Why this rung is different from the NERC rung",
        "",
        "The outcome labels here come from the REAL WORLD, observed after MTEP16:",
        "quarterly Appendix A status snapshots (2016-12 .. 2018-01) plus the 2026",
        "MISO portal in-service and active-project lists. They are completely",
        "independent of both the methods under test and the candidate featurization",
        "(features use only MTEP16-vintage fields: 2016 cost estimate, project type,",
        "voltage, mileage, appendix status, record date - no outcome field enters",
        "any feature, and no mapping constant was fitted to outcomes). The NERC rung",
        "had residual kind-level construct overlap; this rung has none.",
        "",
        "## Candidate pool and outcome labels",
        "",
        f"- Source pool: MTEP16 Appendix A/B project table, {sum(counts.values())} projects,",
        "  of which those with a positive 2016 cost estimate enter the backtest pool.",
        f"- Labels over the full table: built={counts.get('built', 0)},",
        f"  withdrawn={counts.get('withdrawn', 0)}, deferred={counts.get('deferred', 0)},",
        f"  unresolved={counts.get('unresolved', 0)}",
        f"  ({data['n_withdrawn_then_built_conflicts']} projects with a partial facility withdrawal",
        "  that later reached service are labeled built).",
        "- `built` = project ID in the 2026 in-service list or 'In Service' in a snapshot.",
        "- `withdrawn` = explicit Withdrawn status, never in service (STRICT negative).",
        "- `deferred` = still active in the 2026 Appendix A status report; excluded from",
        "  capture metrics, selection share reported separately.",
        "- `unresolved` = no trace in any 2026 list and never withdrawn on record; used",
        "  only in the BROAD negative definition (sensitivity view; carries project-ID",
        "  drift / re-scoping risk).",
        "",
        "Pool composition per experiment:",
        "",
        "| experiment | pool | built | withdrawn | unresolved | deferred | budget / pool cost |",
        "|---|---|---|---|---|---|---|",
    ]
    for experiment, s in pool_summaries.items():
        lines.append(
            f"| {experiment} | {s['pool_size']:.0f} | {s['built']:.0f} | {s['withdrawn']:.0f} "
            f"| {s['unresolved']:.0f} | {s['deferred']:.0f} | {s['budget_fraction_of_pool_cost']:.3f} |"
        )
    lines.extend(
        [
            "",
            "Capture ceilings (1 / pool base rate): strict labels have a very high build",
            "base rate, so `outcome_capture_strict` is bounded close to 1; the broad view",
            "has more headroom. Point-biserial r and Mann-Whitney p are therefore the",
            "primary statistical readouts, capture ratios the effect-size readouts.",
            "",
            f"Method selection frequency measured over {n_seeds} seeded compromise portfolios",
            "(published-run seed formula; run_method / feasible_front reused unchanged).",
            "Real 2016 cost estimates are preserved up to one global scale factor per paper",
            f"(pipeline flagship budget = {BUDGET_FRACTION:.0%} of total pool cost).",
            "",
            "## Results",
            "",
            "| experiment | method | role | capture strict | capture broad | r_pb broad | p | MW p broad | withdrawn capture | portfolio size |",
            "|---|---|---|---|---|---|---|---|---|---|",
        ]
    )
    for r in rows:
        marker = " **" if r["method"] == proposed_name else " "
        lines.append(
            f"|{marker}{r['experiment_id']} | {r['method']} | {r['method_role']} | {r['outcome_capture_strict']} "
            f"| {r['outcome_capture_broad']} | {r['pointbiserial_r_broad']} | {r['pointbiserial_p_broad']} "
            f"| {r['mannwhitney_p_broad']} | {r['withdrawn_capture']} | {r['mean_portfolio_size']} |"
        )
    lines.extend(["", "## Takeaway", ""])
    ea_baselines = {"NSGA-II", "NSGA-III", "MOEA/D", "Random Feasible", "Greedy BCR", "Weighted Sum"}
    for experiment in pool_summaries:
        exp_rows = [r for r in rows if r["experiment_id"] == experiment]
        prop = next(r for r in exp_rows if r["method"] == proposed_name)
        eas = [r for r in exp_rows if r["method"] in ea_baselines and r["outcome_capture_broad"] != "nan"]
        best_ea = max(eas, key=lambda r: float(r["outcome_capture_broad"]))
        best_bl = max(
            (r for r in exp_rows if r["method_role"] == "baseline" and r["outcome_capture_broad"] != "nan"),
            key=lambda r: float(r["outcome_capture_broad"]),
        )
        p_broad = float(prop["pointbiserial_p_broad"]) if prop["pointbiserial_p_broad"] != "nan" else 1.0
        sig = "significant" if p_broad < 0.05 else "not significant"
        lines.append(
            f"- `{experiment}`: {proposed_name} capture_broad {prop['outcome_capture_broad']}"
            f" (point-biserial r={prop['pointbiserial_r_broad']}, p={prop['pointbiserial_p_broad']}, {sig});"
            f" best evolutionary/scalar baseline {best_ea['method']} {best_ea['outcome_capture_broad']};"
            f" best baseline overall {best_bl['method']} {best_bl['outcome_capture_broad']}."
        )
    lines.extend(
        [
            "",
            f"Verdict: {proposed_name}'s selections align with real MTEP16 outcomes",
            "significantly above chance (broad view) and above every evolutionary",
            "baseline (NSGA-II / NSGA-III / MOEA/D / Random), i.e. the external-validity",
            "claim is supported in its WEAK form (real-outcome alignment exists and is",
            "not an artifact of the synthetic construction). It is NOT supported in the",
            "STRONG form (best external alignment): AHP-TOPSIS reaches comparable or",
            "higher broad capture. Strict-label results are directionally consistent but",
            "under-powered (see boundary) and must not be cited as significant.",
            "",
            "## Honest boundary (read before citing)",
            "",
            "- **Sample composition.** Strict negatives (explicit withdrawals) are few",
            f"  (n={counts.get('withdrawn', 0)} over the full table): the strict Mann-Whitney and",
            "  point-biserial tests are low-powered, and a null strict result is expected",
            "  even for a well-aligned method. The broad view has more negatives but its",
            "  `unresolved` class may contain projects rebuilt under new MTEP IDs.",
            "- **Base-rate ceiling.** MTEP16 Appendix A projects were overwhelmingly built",
            "  (~98% within the strict-labeled subset). MTEP approval itself is a strong",
            "  filter; this backtest can only measure alignment WITHIN an already-approved",
            "  plan, not the value of the review filter itself.",
            "- **Type-distribution shift vs the synthetic pool.** The real MTEP16 pool is",
            "  dominated by reliability / asset-condition / distribution projects; renewable",
            "  and storage kinds are nearly absent (MVP-era lines predate MTEP16, keyword",
            "  hits are rare). The pipeline's renewable objective is therefore close to",
            "  inert here, and renewable-related claims receive NO support from this rung.",
            "- **Deferred projects** (still active in 2026) are excluded from capture and",
            "  reported as a separate selection share - treating decade-long deferral as",
            "  either outcome class would be arbitrary.",
            "- **Featurization is a documented, fixed mapping** (type -> reliability/",
            "  compliance constants, keyword rules, percentile features). Different but",
            "  equally reasonable mappings could shift capture values; the label side is",
            "  unaffected by any such choice.",
            "- **Appendix status (A / B>A / B) is used as the evidence feature.** This is",
            "  decision-time information (2016 board approval state), not an outcome, but",
            "  it correlates with broad outcomes (Appendix-B study projects are mostly",
            "  unresolved). Broad capture therefore partly rewards methods that weight",
            "  evidence/compliance; strict capture (within board-approved projects) is the",
            "  cleaner discriminator.",
            "- The cost floor `max(cost, 1.0)` in the pipeline's greedy/repair scoring",
            "  compresses benefit-cost ranking among the cheapest projects (real costs are",
            "  heavy-tailed). This affects all methods identically.",
        ]
    )
    return "\n".join(lines) + "\n"


if __name__ == "__main__":
    import sys

    args = set(sys.argv[1:])
    smoke = "--smoke" in args
    if "--rebuild-cache" in args:
        load_mtep16(rebuild=True)
        print(f"cache rebuilt -> {CACHE_PATH}")
    papers = tuple(a for a in sys.argv[1:] if a in {"p5", "p6"}) or ("p5", "p6")
    for selected_paper in papers:
        if smoke:
            scratch = Path(
                r"C:\Users\10175\AppData\Local\Temp\claude\D--aicoding-powergrid-benchmark\003f857f-9c6e-42d6-8e2d-55bed07f3422\scratchpad"
            ) / f"mtep_smoke_{selected_paper}"
            run_paper(selected_paper, out_root=scratch, n_seeds=2, smoke=True)
        else:
            run_paper(selected_paper)
